from openpyxl import load_workbook

class HandleExcel:
    """
    定义一个处理excel的类
    """
    def __init__(self, filename, sheetname=None):
        self.filename = filename
        self.sheetname = sheetname

    def get_cases(self):
        """
        获取所有测试用例
        :return:
        """
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]
        head_tuple = tuple(ws.iter_rows(max_row=1, values_only=True))[0]
        one_list = []
        for one_tuple in tuple(ws.iter_rows(min_row=2, values_only=True)):
            one_list.append(dict(zip(head_tuple, one_tuple)))
        return one_list

    def get_one_case(self, row):
        """
        获取指定用例
        :return:
        """
        return self.get_cases()[row-1]

    def write_result(self, row,column, result):
        """
        写入结果
        :return:
        """
        wb = load_workbook(self.filename)
        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]

        if isinstance(row, int) and (2 <= row <= ws.max_row):
            ws.cell(row=row, column=column, value=result)
            wb.save(self.filename)
        else:
            print("传入的行号有误，行号应为大于1的整数")


if __name__=="__main__":
    a=('name', 'age')
    b=('wangxin',28)
    zip(a,b)
    c={'wang':34}
    print(dict(c))
    pass