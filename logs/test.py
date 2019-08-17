import openpyxl

class TestHandleExcel:
    """
    测试封装excel类
    :return:
    """

    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def get_cases(self):
        wb = openpyxl.load_workbook(self.filename)

        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]

        list_data = []
        header_data = tuple(ws.iter_rows(max_row=1, values_only=True))[0]

        all_cases = tuple(ws.iter_rows(min_row=2, values_only=True))
        for one_case in all_cases:
            list_data.append(dict(zip(header_data, one_case)))
        return list_data

    def get_one_case(self, row):
        """
        获取单条用例
        :param row:
        :return:
        """
        self.get_cases()[row]

    def write_excel(self, row ,column ,actual_response, result):
        wb = openpyxl.load_workbook(self.filename)

        if self.sheetname is None:
            ws = wb.active
        else:
            ws = wb[self.sheetname]

        if isinstance(row, int) and (2 <= row <= ws.max_row):
            ws.cell(row, column, actual_response)
            wb.save(self.filename)
        else:
            print("传入的行号有误")


wb = openpyxl.load_workbook("wang.xlsx")
# ws = wb.active
ws = wb["deep_account"]
ws.cell(1, 1, value=10)
wb.save("wang.xlsx")
pass
